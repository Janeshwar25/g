import os
import re
import warnings
import requests
import pandas as pd
from openpyxl import load_workbook

from config import Config


def generate_pet_test_script(st_id, template_path, capability_id=None, output_dir='documents'):
    config = Config()

    if not capability_id or not str(capability_id).strip():
        raise ValueError('Capability ID is required.')

    def normalize_text(value):
        if value is None:
            return ''
        return str(value).strip().lower()

    def canonical_metric_key(value):
        # Normalize labels like "Extracts(NF)", "Extracts (NF)", and case variants.
        return re.sub(r'[^a-z0-9]+', '', normalize_text(value))

    def extract_test_cases(description):
        case_install = []

        if 'Case Install LG Fully Insured- OBM' in description:
            case_install.append('Case Install LG Fully Insured- OBM')
        if 'Case Install LG Level Funded- OBM' in description:
            case_install.append('Case Install LG Level Funded- OBM')
        if 'Case Install Manual' in description:
            case_install.append('Case Install Manual')
        if 'Case Install SG Fully Insured- SAMX' in description:
            case_install.append('Case Install SG Fully Insured- SAMX')
        if 'Case Install SG Level Funded- SAMX' in description:
            case_install.append('Case Install SG Level Funded- SAMX')

        fulfillment = []

        if 'Plan Documents' in description:
            fulfillment.append('Plan Documents')
        if 'ID Card' in description:
            fulfillment.append('ID Card')
        if 'NODA' in description:
            fulfillment.append('NODA')
        if 'Benefit Language' in description:
            fulfillment.append('Benefit Language- B360')
        if 'Letters' in description:
            fulfillment.append('Letters')

        extracts = []

        if 'Optum Bank' in description:
            extracts.append('Optum Bank')
        if 'Optum Rx' in description:
            extracts.append('Optum Rx')
        if 'Vision' in description:
            extracts.append('Vision')
        if 'Dental' in description:
            extracts.append('Dental')
        if 'PHAccent' in description:
            extracts.append('PHAccent')
        if 'BOSS' in description:
            extracts.append('BOSS')

        maintenance = []

        if 'Member Group and Member Maintenance- Cirrus' in description:
            maintenance.append('Member Group and Member Maintenance- Cirrus')
        if 'Broker & Employer Portal' in description:
            maintenance.append('Broker & Employer Portal')
        if 'B&amp;E Portal' in description:
            maintenance.append('Broker & Employer Portal')

        bill_comm = []

        if 'Billing' in description:
            bill_comm.append('Billing')
        if 'Commission' in description:
            bill_comm.append('Commission')

        claims_clin = []

        if 'Claims Adjudication' in description:
            claims_clin.append('Claims Adjudication')
        if 'Claims- EOB/PRAs' in description:
            claims_clin.append('Claims- EOB/PRAs')
        if '>Claims<' in description:
            claims_clin.append('Claims- EOB/PRAs')
        if 'ICUE' in description:
            claims_clin.append('Clinical- ICUE')

        pcp = []

        if 'PCP Assignment' in description:
            pcp.append('PCP Assignment')

        digital = []

        if 'myUHC' in description:
            digital.append('myUHC')
        if 'ACET' in description:
            digital.append('ACET')
        if 'Provider Portal' in description:
            digital.append('Provider Portal')
        if 'Beach' in description:
            digital.append('Beach')

        return case_install, fulfillment, extracts, maintenance, bill_comm, claims_clin, pcp, digital

    def get_strategic_theme_ref(theme_id):
        url = f"{config.RALLY_URL}/PortfolioItem/StrategicTheme"
        params = {
            "query": f'(FormattedID = "{theme_id}")',
            "fetch": "_ref,FormattedID,Workspace",
        }

        response = requests.get(
            url,
            headers=config.get_rally_headers(),
            params=params,
            verify=config.VERIFY_SSL,
            timeout=config.REQUEST_TIMEOUT,
        )
        response.raise_for_status()

        result = response.json().get('QueryResult', {}).get('Results', [])
        if result:
            return result[0].get('_ref')
        return None

    def get_rally_data(theme_id):
        ref_st = get_strategic_theme_ref(theme_id)
        if not ref_st:
            return pd.DataFrame()

        url = f"{config.RALLY_URL}/PortfolioItem/SolutionCapability"
        params = {
            "query": f'(Parent = "{ref_st}")',
            "fetch": "_ref,FormattedID,Name,Owner,Description",
            "pageSize": 200,
        }

        response = requests.get(
            url,
            headers=Config.get_rally_headers(),
            params=params,
            verify=config.VERIFY_SSL,
            timeout=config.REQUEST_TIMEOUT,
        )
        response.raise_for_status()

        result = response.json().get('QueryResult', {}).get('Results', [])
        rows = []
        for capability in result:
            rows.append({
                'Solution Capability': capability.get('FormattedID'),
                'Owner': capability.get('Owner', {}).get('_refObjectName'),
                'Name': capability.get('Name'),
                'Description': capability.get('Description'),
            })

        return pd.DataFrame(rows)

    def filter_test_plan_workbook(output_path, sheet_filters, filter_column_contains='Application/Component'):
        warnings.filterwarnings(
            'ignore',
            message='Data Validation extension is not supported and will be removed',
            category=UserWarning,
            module='openpyxl.worksheet._reader',
        )

        workbook = load_workbook(template_path)
        mapping_sheet_names = list(sheet_filters.keys())

        for sheet_name, filter_values in sheet_filters.items():
            if sheet_name not in workbook.sheetnames:
                continue

            worksheet = workbook[sheet_name]

            if not filter_values:
                workbook.remove(worksheet)
                continue

            header_cells = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]
            filter_col_idx = None

            for idx, cell in enumerate(header_cells, start=1):
                header_value = str(cell.value or '').replace('\n', ' ').strip().lower()
                if filter_column_contains.strip().lower() in header_value:
                    filter_col_idx = idx
                    break

            if filter_col_idx is None:
                continue

            allowed_values = {normalize_text(item) for item in filter_values}

            rows_to_delete = []
            for row_idx, (cell_value,) in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    max_row=worksheet.max_row,
                    min_col=filter_col_idx,
                    max_col=filter_col_idx,
                    values_only=True,
                ),
                start=2,
            ):
                if normalize_text(cell_value) not in allowed_values:
                    rows_to_delete.append(row_idx)

            if rows_to_delete:
                blocks = []
                block_start = rows_to_delete[0]
                block_end = rows_to_delete[0]

                for row_idx in rows_to_delete[1:]:
                    if row_idx == block_end + 1:
                        block_end = row_idx
                    else:
                        blocks.append((block_start, block_end))
                        block_start = row_idx
                        block_end = row_idx
                blocks.append((block_start, block_end))

                for block_start, block_end in reversed(blocks):
                    worksheet.delete_rows(block_start, block_end - block_start + 1)

        # Reconcile Metrics rows based on the workbook's final state after all filtering.
        # Keep metrics only for mapped sheets that still exist and still have data rows.
        if 'Metrics' in workbook.sheetnames:
            metrics_sheet = workbook['Metrics']
            sheet_names_to_keep = {
                sheet_name
                for sheet_name in mapping_sheet_names
                if sheet_name in workbook.sheetnames and workbook[sheet_name].max_row > 1
            }
            canonical_keep_names = {canonical_metric_key(name) for name in sheet_names_to_keep}
            canonical_mapped_names = {canonical_metric_key(name) for name in mapping_sheet_names}

            metric_rows_to_delete = []
            for row_idx, (cell_value,) in enumerate(
                metrics_sheet.iter_rows(
                    min_row=1,
                    max_row=metrics_sheet.max_row,
                    min_col=2,
                    max_col=2,
                    values_only=True,
                ),
                start=1,
            ):
                canonical_value = canonical_metric_key(cell_value)
                # Only act on rows that correspond to mapped sheets.
                if canonical_value in canonical_mapped_names and canonical_value not in canonical_keep_names:
                    metric_rows_to_delete.append(row_idx)

            for row_idx in reversed(metric_rows_to_delete):
                metrics_sheet.delete_rows(row_idx, 1)

        if len(workbook.sheetnames) == 0:
            placeholder_sheet = workbook.create_sheet('No Test Cases')
            placeholder_sheet['A1'] = 'No test cases were available to retain.'

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        workbook.save(output_path)
        return output_path

    strategic_theme = st_id.strip().upper()
    capability = str(capability_id).strip().upper()
    rally_df = get_rally_data(strategic_theme)
    if rally_df.empty:
        raise ValueError('No capabilities found for this Strategic Theme.')

    selected_capability_df = rally_df[
        rally_df['Solution Capability'].fillna('').str.strip().str.upper() == capability
    ]
    if selected_capability_df.empty:
        raise ValueError('Capability ID not found under this Strategic Theme.')

    description = ' '.join(selected_capability_df['Description'].dropna().astype(str).tolist())
    if not description.strip():
        raise ValueError('No description content found in selected capability.')

    case_install, fulfillment, extracts, maintenance, bill_comm, claims_clin, pcp, digital = extract_test_cases(description)

    sheet_filters = {
        'Case install': case_install,
        'Fulfillment': fulfillment,
        'Extracts(NF)': extracts,
        'Maintenance': maintenance,
        'Bill&Comm': bill_comm,
        'ClmsClin': claims_clin,
        'PCP Assignment': pcp,
        'Digital': digital,
    }

    output_filename = f"{strategic_theme}_NB_PET_Test_Script.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    generated_path = filter_test_plan_workbook(
        output_path=output_path,
        sheet_filters=sheet_filters,
        filter_column_contains='Application/Component',
    )

    return generated_path

# import json
# import pandas as pd
# from engine.utils import read_excel_to_dataframe, filter_rows_by_value
# import os
# import requests
# import warnings
# from pprint import pprint 
# from openpyxl import load_workbook

# rally_url = 'https://rally1.rallydev.com/slm/webservice/v2.0'

# WORKSPACE = 'UHG'
# PROJECT = 'Pioneers GenAI'

# RALLY_KEY='_DRLmbCUhTiGNpD5pIR3oAkoB2ypaE96pWDCGK6cfZA'

# HEADERS = {
#     'ZSESSIONID': RALLY_KEY,
#     'Content-Type': 'application/json'
# }

# def get_strategic_theme_ref(st_id):
#         url = f"{rally_url}/PortfolioItem/StrategicTheme"
#         params = {
#             "query": f'(FormattedID = "{st_id}")',
#             "fetch": "_ref,FormattedID,Workspace",
#         }

#         response = requests.get(url, headers=HEADERS, params=params, verify=False).json()
#         result = response["QueryResult"]["Results"]
#         if result:
#             return result[0]["_ref"]
#         return None

# def get_rally_data(rally_st):
#     ref_st = get_strategic_theme_ref(rally_st)
#     url = f"{rally_url}/PortfolioItem/SolutionCapability"
#     params = {
#         "query": f'(Parent = "{ref_st}")',
#         "fetch": "_ref,FormattedID,Name,Owner,Description",
#         "pageSize": 200,
#     }

#     response = requests.get(url, headers=HEADERS, params=params, verify=False)

#     if response.status_code == 200:
#         data = response.json()
#         result = data['QueryResult']['Results']
#         rows = []
#         for c in result:
#             cap_details = {
#                 'Solution Capability': c["FormattedID"],
#                 'Owner': c.get('Owner', {}).get('_refObjectName', None),
#                 'Name': c.get('Name', None),
#                 'Description': c.get('Description', None)
#             }
#             rows.append(cap_details)
#         df = pd.DataFrame(rows)
#         return df
#     return []

# x = get_rally_data('ST20983')
# print(x.columns)
# print(x)

# x_testing = x[x['Name'].str.contains('Testing', na=False) & x['Name'].str.contains('PET', na=False)]

# print(x_testing)

# description = x_testing.iloc[0]['Description']

# print(description)

# def extract_test_cases(description):
#     case_install = []

#     if 'Case Install LG Fully Insured- OBM' in description:
#         case_install.append('Case Install LG Fully Insured- OBM')
#     if 'Case Install LG Level Funded- OBM' in description:
#         case_install.append('Case Install LG Level Funded- OBM')
#     if 'Case Install Manual' in description:
#         case_install.append('Case Install Manual')
#     if 'Case Install SG Fully Insured- SAMX' in description:
#         case_install.append('Case Install SG Fully Insured- SAMX')
#     if 'Case Install SG Level Funded- SAMX' in description:
#         case_install.append('Case Install SG Level Funded- SAMX')

#     fulfillment = []
    
#     if 'Plan Documents' in description:
#         fulfillment.append('Plan Documents')
#     if 'ID Card' in description:
#         fulfillment.append('ID Card')
#     if 'NODA' in description:
#         fulfillment.append('NODA')
#     if 'Benefit Language' in description:
#         fulfillment.append('Benefit Language- B360')
#     if 'Letters' in description:
#         fulfillment.append('Letters')

#     extracts = []

#     if 'Optum Bank' in description:
#         extracts.append('Optum Bank')
#     if 'Optum Rx' in description:
#         extracts.append('Optum Rx')
#     if 'Vision' in description:
#         extracts.append('Vision')
#     if 'Dental' in description:
#         extracts.append('Dental')
#     if 'PHAccent' in description:
#         extracts.append('PHAccent')
#     if 'BOSS' in description:
#         extracts.append('BOSS')
    
#     maintenance = []

#     if 'Member Group and Member Maintenance- Cirrus' in description:
#         maintenance.append('Member Group and Member Maintenance- Cirrus')
#     if 'Broker & Employer Portal' in description:
#         maintenance.append('Broker & Employer Portal')
#     if 'B&amp;E Portal' in description:
#         maintenance.append('Broker & Employer Portal')

#     bill_comm = []

#     if 'Billing' in description:
#         bill_comm.append('Billing')
#     if 'Commission' in description:
#         bill_comm.append('Commission')
    
#     claims_clin = []

#     if 'Claims Adjudication' in description:
#         claims_clin.append('Claims Adjudication')
#     if 'Claims- EOB/PRAs' in description:
#         claims_clin.append('Claims- EOB/PRAs')
#     if '>Claims<' in description:
#         claims_clin.append('Claims- EOB/PRAs')
#     if 'ICUE' in description:
#         claims_clin.append('Clinical- ICUE')
    
#     pcp = []

#     if 'PCP Assignment' in description:
#         pcp.append('PCP Assignment')
    
#     digital = []

#     if 'myUHC' in description:
#         digital.append('myUHC')
#     if 'ACET' in description:
#         digital.append('ACET')
#     if 'Provider Portal' in description:
#         digital.append('Provider Portal')
#     if 'Beach' in description:
#         digital.append('Beach')

#     return case_install, fulfillment, extracts, maintenance, bill_comm, claims_clin, pcp, digital


# def normalize_text(value):
#     if value is None:
#         return ''
#     return str(value).strip().lower()


# def filter_test_plan_workbook(template_path, output_path, sheet_filters, filter_column_contains='Application/Component'):
#     warnings.filterwarnings(
#         'ignore',
#         message='Data Validation extension is not supported and will be removed',
#         category=UserWarning,
#         module='openpyxl.worksheet._reader',
#     )

#     workbook = load_workbook(template_path)

#     for sheet_name, filter_values in sheet_filters.items():
#         if sheet_name not in workbook.sheetnames:
#             print(f"Sheet not found, skipping: {sheet_name}")
#             continue

#         worksheet = workbook[sheet_name]

#         if not filter_values:
#             workbook.remove(worksheet)
#             print(f"Removed sheet (no test cases): {sheet_name}")
#             continue

#         header_cells = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]
#         filter_col_idx = None

#         for idx, cell in enumerate(header_cells, start=1):
#             header_value = str(cell.value or '').replace('\n', ' ').strip().lower()
#             if filter_column_contains.strip().lower() in header_value:
#                 filter_col_idx = idx
#                 break

#         if filter_col_idx is None:
#             print(f"Column containing '{filter_column_contains}' not found in sheet: {sheet_name}. Skipping filter.")
#             continue

#         allowed_values = {normalize_text(item) for item in filter_values}

#         rows_to_delete = []
#         for row_idx, (cell_value,) in enumerate(
#             worksheet.iter_rows(
#                 min_row=2,
#                 max_row=worksheet.max_row,
#                 min_col=filter_col_idx,
#                 max_col=filter_col_idx,
#                 values_only=True,
#             ),
#             start=2,
#         ):
#             if normalize_text(cell_value) not in allowed_values:
#                 rows_to_delete.append(row_idx)

#         if rows_to_delete:
#             blocks = []
#             block_start = rows_to_delete[0]
#             block_end = rows_to_delete[0]

#             for row_idx in rows_to_delete[1:]:
#                 if row_idx == block_end + 1:
#                     block_end = row_idx
#                 else:
#                     blocks.append((block_start, block_end))
#                     block_start = row_idx
#                     block_end = row_idx
#             blocks.append((block_start, block_end))

#             for block_start, block_end in reversed(blocks):
#                 worksheet.delete_rows(block_start, block_end - block_start + 1)

#         print(f"Filtered sheet: {sheet_name} | kept values: {filter_values}")

#     if len(workbook.sheetnames) == 0:
#         placeholder_sheet = workbook.create_sheet('No Test Cases')
#         placeholder_sheet['A1'] = 'No test cases were available to retain.'

#     os.makedirs(os.path.dirname(output_path), exist_ok=True)
#     workbook.save(output_path)
#     return output_path

# case_install, fulfillment, extracts, maintenance, bill_comm, claims_clin, pcp, digital = extract_test_cases(description)

# template_path = 'documents/Template_NB PET_Test Scripts.xlsx'
# output_path = 'documents/Filtered_NB_PET_Test_Scripts.xlsx'

# sheet_filters = {
#     'Case install': case_install,
#     'Fulfillment': fulfillment,
#     'Extracts(NF)': extracts,
#     'Maintenance': maintenance,
#     'Bill&Comm': bill_comm,
#     'ClmsClin': claims_clin,
#     'PCP Assignment': pcp,
#     'Digital': digital,
# }

# saved_file = filter_test_plan_workbook(
#     template_path=template_path,
#     output_path=output_path,
#     sheet_filters=sheet_filters,
#     filter_column_contains='Application/Component',
# )

# print(f"\nFiltered workbook created: {saved_file}")

# print("Case Install Test Cases:")
# print(case_install)

# print("\nFulfillment Test Cases:")
# print(fulfillment)

# print("\nExtracts Test Cases:")
# print(extracts)

# print("\nMaintenance Test Cases:")
# print(maintenance)

# print("\nBilling & Commission Test Cases:")
# print(bill_comm)

# print("\nClaims & Clinical Test Cases:")
# print(claims_clin)

# print("\nPCP Test Cases:")
# print(pcp)

# print("\nDigital Test Cases:")
# print(digital)

# test_plan = pd.read_excel('documents/Template_NB PET_Test Scripts.xlsx')


